from __future__ import annotations

import csv
import hashlib
import json
import re
from collections.abc import Iterable
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import urlparse

import httpx
from pdfminer.high_level import extract_text

from app.core.config import settings
from app.ingestion.column_mapping import ColumnMapping, ColumnMappingService
from app.ingestion.deployment_extractor import (
    deployment_tables_from_structured_tables,
    json_deployment_tables,
    markdown_tables,
    table_metadata,
    yaml_deployment_tables,
)
from app.ingestion.github_source_resolver import resolve_github_source
from app.ingestion.html_table_extractor import extract_html_deployment_tables
from app.ingestion.intake_models import CandidatePreview, ParsedSource, SourceArtifact, SourceFingerprint
from app.ingestion.network_normalizer import NetworkNormalizer
from app.ingestion.solidity_address_extractor import extract_solidity_deployment_table


ADDRESS_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:0x[a-fA-F0-9]{40,64}|bc1[ac-hj-np-z02-9]{11,87}|[13][a-km-zA-HJ-NP-Z1-9]{25,34}|T[1-9A-HJ-NP-Za-km-z]{33}|r[1-9A-HJ-NP-Za-km-z]{24,34}|(?:EQ|UQ)[A-Za-z0-9_-]{46})(?![A-Za-z0-9])",
    re.IGNORECASE,
)
URL_RE = re.compile(r"^https?://", re.IGNORECASE)
XRP_RE = re.compile(r"^r[1-9A-HJ-NP-Za-km-z]{24,34}$")
EVM_RE = re.compile(r"^0x[a-fA-F0-9]{40}$")
LONG_0X_RE = re.compile(r"^0x[a-fA-F0-9]{40,64}$")
BTC_RE = re.compile(r"^(?:bc1[ac-hj-np-z02-9]{11,87}|[13][a-km-zA-HJ-NP-Z1-9]{25,34})$", re.IGNORECASE)
LTC_RE = re.compile(r"^(?:ltc1[ac-hj-np-z02-9]{11,87}|[LM3][a-km-zA-HJ-NP-Z1-9]{25,34})$", re.IGNORECASE)
TRON_RE = re.compile(r"^T[1-9A-HJ-NP-Za-km-z]{33}$")
TON_RE = re.compile(r"^(?:EQ|UQ)[A-Za-z0-9_-]{46}$")
SUBSTRATE_RE = re.compile(r"^[1-9A-HJ-NP-Za-km-z]{32,64}$")
COSMOS_RE = re.compile(r"^cosmos1[0-9a-z]{20,80}$", re.IGNORECASE)
DYDX_RE = re.compile(r"^dydx1[0-9a-z]{20,80}$", re.IGNORECASE)
DOGE_RE = re.compile(r"^D[1-9A-HJ-NP-Za-km-z]{25,34}$")
AVALANCHE_X_RE = re.compile(r"^X-avax1[0-9a-z]{20,80}$", re.IGNORECASE)
SOLANA_RE = re.compile(r"^[1-9A-HJ-NP-Za-km-z]{32,44}$")
HEDERA_RE = re.compile(r"^0\.0\.\d{1,12}$")
VAULTA_RE = re.compile(r"^[a-z1-5.]{3,12}$")
XDC_RE = re.compile(r"^xdc[a-fA-F0-9]{40}$")
ALGORAND_RE = re.compile(r"^[A-Z2-7]{58}$")
NEAR_RE = re.compile(r"^(?:[a-fA-F0-9]{64}|[a-z0-9_-]+(?:\.[a-z0-9_-]+)+)$")
NOBLE_RE = re.compile(r"^noble1[0-9a-z]{20,80}$", re.IGNORECASE)
TEZOS_RE = re.compile(r"^tz1[1-9A-HJ-NP-Za-km-z]{30,40}$")
HACKEN_STOP_MARKERS = {"collateral ratios", "team composition", "conclusion", "disclaimers"}
HACKEN_TOC_MARKERS = {
    "executive summary",
    "building trust",
    "methodology",
    "proof of reserves scope findings",
    "proof of reserves scope and findings",
    "team composition",
    "conclusion",
    "disclaimers",
    "references",
}
PDF_LIGATURE_TRANSLATION = str.maketrans(
    {
        "ﬀ": "ff",
        "ﬁ": "fi",
        "ﬂ": "fl",
        "ﬃ": "ffi",
        "ﬄ": "ffl",
    }
)
PDF_TEXT_REPLACEMENTS = (
    ("\ufb00", "ff"),
    ("\ufb01", "fi"),
    ("\ufb02", "fl"),
    ("\ufb03", "ffi"),
    ("\ufb04", "ffl"),
    ("\u00ef\u00ac\u20ac", "ff"),
    ("\u00ef\u00ac\u0081", "fi"),
    ("\u00ef\u00ac\u201a", "fl"),
    ("\u00ef\u00ac\u0192", "ffi"),
    ("\u00ef\u00ac\u201e", "ffl"),
)
HACKEN_NETWORKS = sorted(
    {
        "Algorand",
        "Aptos",
        "Arbitrum",
        "Arbitrum Nova",
        "Arbitrum One",
        "Aurora",
        "Avalanche",
        "Avalanche-C",
        "Avalanche C-Chain",
        "Avalanche C Chain",
        "Avalanche-X",
        "Base",
        "Bera",
        "Blast",
        "BLAST",
        "BItcoin",
        "Bitcoin",
        "BSC",
        "Celo",
        "Codex",
        "Corn",
        "Cosmos",
        "Dogecoin",
        "DYDX",
        "Ethereum",
        "Fantom",
        "Hedera",
        "Hyperliquid",
        "HyperEVM",
        "Kaia",
        "Kava EVM",
        "KCC",
        "Linea",
        "Litecoin",
        "Manta",
        "Mantle",
        "Merlin",
        "Monad",
        "Morph",
        "NEAR",
        "Near",
        "Noble",
        "Optimism",
        "Plasma",
        "Polkadot AH",
        "Polygon",
        "Ripple",
        "XRP Ledger",
        "Scroll",
        "SEI",
        "Sei EVM",
        "Solana",
        "Sonic",
        "Starknet",
        "Statemint",
        "Sui",
        "SUI",
        "Taiko",
        "Tezos",
        "Ton",
        "TON",
        "Tron",
        "Unichain",
        "Vaulta",
        "XDC",
        "Zircuit",
        "ZKLink Nova",
        "zkLink Nova",
        "ZKSync Era",
        "zkSync Era",
        "ZKSync Lite",
        "zkSync Lite",
    },
    key=lambda value: len(value.split()),
    reverse=True,
)
HACKEN_NETWORK_ALIASES = {network: network for network in HACKEN_NETWORKS}
HACKEN_NETWORK_ALIASES.update(
    {
        "Avalanche": "Avalanche-C",
        "Avalanche C-Chain": "Avalanche-C",
        "Avalanche C Chain": "Avalanche-C",
        "BItcoin": "Bitcoin",
        "BLAST": "Blast",
        "NEAR": "Near",
        "SUI": "Sui",
        "TON": "Ton",
        "ZKLink Nova": "zkLink Nova",
        "ZKSync Era": "zkSync Era",
        "ZKSync Lite": "zkSync Lite",
    }
)
HACKEN_NETWORK_LABELS = sorted(HACKEN_NETWORK_ALIASES, key=lambda value: len(value.split()), reverse=True)
COMPACT_HACKEN_NETWORK_ALIASES = {
    "Algorand": "Algorand",
    "Aptos": "Aptos",
    "Arbitrum": "Arbitrum",
    "ArbitrumNova": "Arbitrum Nova",
    "ArbitrumOne": "Arbitrum One",
    "Aurora": "Aurora",
    "Avalanche-C": "Avalanche-C",
    "AvalancheC": "Avalanche-C",
    "AvalancheC-Chain": "Avalanche-C",
    "AvalancheCChain": "Avalanche-C",
    "Avalanche-X": "Avalanche-X",
    "AvalancheX": "Avalanche-X",
    "Base": "Base",
    "Bera": "Bera",
    "Blast": "Blast",
    "BLAST": "Blast",
    "BItcoin": "Bitcoin",
    "Bitcoin": "Bitcoin",
    "BSC": "BSC",
    "Celo": "Celo",
    "Codex": "Codex",
    "Corn": "Corn",
    "Cosmos": "Cosmos",
    "Dogecoin": "Dogecoin",
    "DYDX": "DYDX",
    "Ethereum": "Ethereum",
    "Hedera": "Hedera",
    "Hyperliquid": "Hyperliquid",
    "HyperEVM": "HyperEVM",
    "Kaia": "Kaia",
    "KavaEVM": "Kava EVM",
    "KCC": "KCC",
    "Linea": "Linea",
    "Litecoin": "Litecoin",
    "Manta": "Manta",
    "Mantle": "Mantle",
    "Merlin": "Merlin",
    "Monad": "Monad",
    "Morph": "Morph",
    "NEAR": "Near",
    "Near": "Near",
    "Noble": "Noble",
    "Optimism": "Optimism",
    "Plasma": "Plasma",
    "PolkadotAH": "Polkadot AH",
    "Polygon": "Polygon",
    "Ripple": "Ripple",
    "XRPLedger": "XRP Ledger",
    "Scroll": "Scroll",
    "SEI": "SEI",
    "SeiEVM": "Sei EVM",
    "Solana": "Solana",
    "Sonic": "Sonic",
    "Starknet": "Starknet",
    "Statemint": "Statemint",
    "Sui": "Sui",
    "SUI": "Sui",
    "Taiko": "Taiko",
    "Tezos": "Tezos",
    "Ton": "Ton",
    "TON": "Ton",
    "Tron": "Tron",
    "Unichain": "Unichain",
    "Vaulta": "Vaulta",
    "XDC": "XDC",
    "Zircuit": "Zircuit",
    "ZKLinkNova": "zkLink Nova",
    "zkLinkNova": "zkLink Nova",
    "ZKSyncEra": "ZKSync Era",
    "zkSyncEra": "zkSync Era",
    "ZKSyncLite": "ZKSync Lite",
    "zkSyncLite": "zkSync Lite",
}
COMPACT_HACKEN_NETWORKS = sorted(
    {re.sub(r"\s+", "", alias): canonical for alias, canonical in COMPACT_HACKEN_NETWORK_ALIASES.items()}.items(),
    key=lambda item: len(item[0]),
    reverse=True,
)
COMPACT_HACKEN_STOP_MARKERS = [
    "collateral ratios",
    "team composition",
    "conclusion",
    "disclaimers",
]
CONTROL_SHEET_NAMES = {
    "summary",
    "schema",
    "source provenance",
    "por search backlog",
    "indodax cmc capture status",
    "capture status",
    "cmc capture status",
    "search backlog",
    "backlog",
    "provenance",
    "readme",
    "read me",
    "notes",
    "metadata",
    "instructions",
    "config",
    "settings",
    "toc",
    "cover",
}
SHEET_ENTITY_SUFFIX_TOKENS = {"wallets", "wallet", "staking", "validators", "validator", "addresses", "address", "cmc", "eth", "cet"}
SHEET_ENTITY_CASING = {
    "okx": "OKX",
    "mexc": "MEXC",
    "kucoin": "KuCoin",
    "bybit": "Bybit",
    "bitfinex": "Bitfinex",
    "bitmex": "BitMEX",
    "huobi": "Huobi",
    "htx": "HTX",
    "coinex": "CoinEx",
    "indodax": "Indodax",
    "deribit": "Deribit",
    "firi": "Firi",
}


class SourceAdapter:
    adapter_name = "base_adapter"

    def parse(self, artifact: SourceArtifact, fingerprint: SourceFingerprint, raw_content: bytes) -> ParsedSource:
        text = _decode(raw_content, fallback=artifact.pasted_text or "")
        table_preview = _markdown_tables(text)
        candidates = _extract_text_candidates(artifact, fingerprint, text)
        return _parsed(
            artifact,
            fingerprint,
            document_text=text,
            document_title=None,
            metadata={"source_input_type": "plain_text"},
            table_preview=table_preview,
            candidates=candidates,
        )


class PlainTextAdapter(SourceAdapter):
    adapter_name = "plain_text_adapter"


class JsonYamlAdapter(SourceAdapter):
    adapter_name = "json_yaml_adapter"

    def parse(self, artifact: SourceArtifact, fingerprint: SourceFingerprint, raw_content: bytes) -> ParsedSource:
        text = _decode(raw_content, fallback=artifact.pasted_text or "")
        tables = json_deployment_tables(
            text,
            source_url=artifact.source_url,
            source_input_type="json_deployment_registry",
            evidence_type="official_docs_deployment" if artifact.source_url else "source_extraction_context",
        )
        if not tables:
            tables = yaml_deployment_tables(
                text,
                source_url=artifact.source_url,
                source_input_type="yaml_deployment_registry",
                evidence_type="official_docs_deployment" if artifact.source_url else "source_extraction_context",
            )
        if tables:
            source_input_type = (tables[0].get("metadata") or {}).get("source_input_type") or "structured_deployment_registry"
            candidates = _extract_table_candidates(artifact, fingerprint, tables, default_source_input_type=source_input_type)
            metadata = table_metadata(tables, source_input_type=source_input_type, text=text, source_url=artifact.source_url)
            return _parsed(
                artifact,
                fingerprint,
                document_text=text,
                document_title=_title_from_url(artifact.source_url),
                metadata=metadata,
                table_preview=tables,
                candidates=candidates,
            )
        candidates = _extract_text_candidates(artifact, fingerprint, text, source_input_type="structured_text_registry")
        metadata = {"source_input_type": "structured_text_registry"}
        return _parsed(artifact, fingerprint, document_text=text, document_title=None, metadata=metadata, table_preview=[], candidates=candidates)


class GitHubAdapter(SourceAdapter):
    adapter_name = "github_adapter"

    def parse(self, artifact: SourceArtifact, fingerprint: SourceFingerprint, raw_content: bytes) -> ParsedSource:
        text, resolved_url = resolve_github_source(artifact.source_url, raw_content)
        if not text:
            text = artifact.pasted_text or ""
        if not text and artifact.source_url:
            text = f"GitHub source: {artifact.source_url}"

        tables = _github_deployment_tables(text, source_url=resolved_url or artifact.source_url)
        if tables:
            source_input_type = (tables[0].get("metadata") or {}).get("source_input_type") or "official_github_deployment_table"
            candidates = _extract_table_candidates(artifact, fingerprint, tables, default_source_input_type=source_input_type)
            metadata = table_metadata(tables, source_input_type=source_input_type, text=text, source_url=resolved_url or artifact.source_url)
            metadata["resolved_source_url"] = resolved_url
            return _parsed(
                artifact,
                fingerprint,
                document_text=text,
                document_title=_title_from_url(resolved_url or artifact.source_url),
                metadata=metadata,
                table_preview=tables,
                candidates=candidates,
            )

        candidates = _extract_text_candidates(artifact, fingerprint, text, source_input_type="github_source")
        metadata = {"source_input_type": fingerprint.final_source_type or "github_source", "resolved_source_url": resolved_url}
        return _parsed(artifact, fingerprint, document_text=text, document_title=_title_from_url(resolved_url or artifact.source_url), metadata=metadata, table_preview=[], candidates=candidates)


class WebDocsAdapter(SourceAdapter):
    adapter_name = "web_docs_adapter"

    def parse(self, artifact: SourceArtifact, fingerprint: SourceFingerprint, raw_content: bytes) -> ParsedSource:
        text = _decode(raw_content, fallback=artifact.pasted_text or "")
        tables = _web_docs_deployment_tables(text, source_url=artifact.source_url, content_type=artifact.content_type)
        warnings: list[str] = []
        if tables:
            source_input_type = (tables[0].get("metadata") or {}).get("source_input_type") or "docs_deployment_table"
            candidates = _extract_table_candidates(artifact, fingerprint, tables, default_source_input_type=source_input_type)
            metadata = table_metadata(tables, source_input_type=source_input_type, text=text, source_url=artifact.source_url)
        else:
            candidates = []
            warnings.append("docs_table_not_detected")
            metadata = {"source_input_type": "docs_text", "table_count": 0, "warnings": warnings}
        return _parsed(
            artifact,
            fingerprint,
            document_text=text,
            document_title=_title_from_url(artifact.source_url),
            metadata=metadata,
            table_preview=tables,
            candidates=candidates,
            warnings=warnings,
        )


class PdfAdapter(SourceAdapter):
    adapter_name = "pdf_adapter"

    def parse(self, artifact: SourceArtifact, fingerprint: SourceFingerprint, raw_content: bytes) -> ParsedSource:
        warnings: list[str] = []
        text = ""
        if raw_content:
            try:
                if settings.pdf_max_pages and settings.pdf_max_pages > 0:
                    text = extract_text(BytesIO(raw_content), maxpages=settings.pdf_max_pages)
                else:
                    text = extract_text(BytesIO(raw_content))
            except Exception:
                text = raw_content[:16_000].decode("utf-8", errors="ignore")
                warnings.append("pdf_text_extraction_fallback")
        normalized_text = _normalize_pdf_text(text)
        text_normalized = normalized_text != text
        text = normalized_text
        entity_name = _detect_pdf_entity(text)
        table_preview = _parse_hacken_layout_wallet_rows(raw_content, text, entity_name) if raw_content and _is_hacken_por_text(text, entity_name) else []
        if not table_preview:
            table_preview = _parse_hacken_audited_wallet_rows(text)
        diagnostics = _pdf_audited_wallet_diagnostics(text, table_preview)
        if table_preview:
            candidates = _extract_table_candidates(artifact, fingerprint, table_preview, default_source_input_type="pdf_audited_wallet_table")
            diagnostics["pdf_parser_mode"] = _hacken_pdf_parser_mode(table_preview)
            diagnostics.update(_hacken_wallet_preview_metadata(table_preview, candidates))
            metadata = {
                "source_input_type": "pdf_audited_wallet_table",
                "entity_name": entity_name,
                "category": "cex",
                "sub_category": "reserve_boundary",
                "expected_roles": ["cex_por_wallet"],
                "table_count": 1,
                "pdf_text_normalized": text_normalized,
                "warnings": warnings,
                **diagnostics,
            }
        else:
            hacken_por = _is_hacken_por_text(text, entity_name)
            if hacken_por:
                warnings.append("pdf_structured_hacken_parser_failed")
                candidates = []
            else:
                if diagnostics["audited_wallet_heading_found"]:
                    warnings.append("pdf_audited_wallet_section_found_but_no_rows")
                warnings.append("pdf_loose_text_fallback_used")
                candidates = _extract_text_candidates(artifact, fingerprint, text, source_input_type="pdf_text_fallback")
            diagnostics["pdf_parser_mode"] = "pdf_text_fallback"
            metadata = {
                "source_input_type": "pdf_text_fallback",
                "entity_name": entity_name,
                "pdf_text_normalized": text_normalized,
                "warnings": warnings,
                **diagnostics,
            }
            if hacken_por or (entity_name in {"Bybit", "KuCoin", "MEXC"} and _is_proof_of_reserves_text(text)):
                metadata.update(
                    {
                        "category": "cex",
                        "sub_category": "reserve_boundary",
                        "expected_roles": ["cex_por_wallet"],
                    }
                )
        return _parsed(artifact, fingerprint, document_text=text, document_title=_first_line(text), metadata=metadata, table_preview=table_preview, candidates=candidates, warnings=warnings)


class ExcelCsvAdapter(SourceAdapter):
    adapter_name = "excel_csv_adapter"

    def parse(self, artifact: SourceArtifact, fingerprint: SourceFingerprint, raw_content: bytes) -> ParsedSource:
        if fingerprint.final_source_type == "csv_upload":
            tables = _csv_tables(raw_content)
            metadata = {
                "source_input_type": "csv_registry",
                "sheet_count": 0,
                "parsed_sheet_names": [],
                "skipped_sheet_names": [],
            }
        else:
            tables, parsed_sheets, skipped_sheets, warnings = _xlsx_tables(raw_content)
            metadata = {
                "source_input_type": "xlsx_multi_sheet_registry" if len(parsed_sheets) > 1 else "xlsx_registry",
                "sheet_count": len(parsed_sheets) + len(skipped_sheets),
                "parsed_sheet_names": parsed_sheets,
                "skipped_sheet_names": skipped_sheets,
                "warnings": warnings,
            }
        candidates = _extract_table_candidates(artifact, fingerprint, tables, default_source_input_type=metadata["source_input_type"])
        table_text = "\n".join(_tables_to_lines(tables))
        return _parsed(
            artifact,
            fingerprint,
            document_text=table_text,
            document_title=_file_name(artifact.local_file_path) or artifact.filename,
            metadata={**metadata, "table_count": len(tables)},
            table_preview=tables,
            candidates=candidates,
            warnings=metadata.get("warnings", []),
        )


ADAPTERS = {
    "plain_text_adapter": PlainTextAdapter,
    "json_yaml_adapter": JsonYamlAdapter,
    "github_adapter": GitHubAdapter,
    "web_docs_adapter": WebDocsAdapter,
    "pdf_adapter": PdfAdapter,
    "excel_csv_adapter": ExcelCsvAdapter,
}


def adapter_by_name(adapter_name: str) -> SourceAdapter:
    try:
        return ADAPTERS[adapter_name]()
    except KeyError as exc:
        raise ValueError(f"Unknown adapter: {adapter_name}") from exc


def _github_deployment_tables(text: str, *, source_url: str | None) -> list[dict]:
    path = urlparse(source_url or "").path.lower()
    if path.endswith(".sol") or _looks_like_solidity(text):
        tables = extract_solidity_deployment_table(text, source_url=source_url)
        if tables:
            return tables
    if path.endswith((".json", ".jsonc")) or text.lstrip().startswith(("{", "[")):
        tables = json_deployment_tables(
            text,
            source_url=source_url,
            source_input_type="github_json_deployment_registry",
            evidence_type="official_github_deployment",
        )
        if tables:
            return tables
    if path.endswith((".yaml", ".yml")):
        tables = yaml_deployment_tables(
            text,
            source_url=source_url,
            source_input_type="github_json_deployment_registry",
            evidence_type="official_github_deployment",
        )
        if tables:
            return tables
    if _looks_like_html(text):
        tables = extract_html_deployment_tables(text, source_url=source_url)
        if tables:
            return _retag_tables(tables, source_input_type="github_markdown_deployment_table", evidence_type="official_github_deployment")
    if "|" in text:
        tables = deployment_tables_from_structured_tables(
            markdown_tables(text),
            source_url=source_url,
            source_input_type="github_markdown_deployment_table",
            evidence_type="official_github_deployment",
            text=text,
        )
        if tables:
            return tables
    return []


def _web_docs_deployment_tables(text: str, *, source_url: str | None, content_type: str | None) -> list[dict]:
    if _looks_like_html(text) or "html" in (content_type or "").lower():
        tables = extract_html_deployment_tables(text, source_url=source_url)
        if tables:
            return tables
    if "|" in text:
        tables = deployment_tables_from_structured_tables(
            markdown_tables(text),
            source_url=source_url,
            source_input_type="docs_markdown_deployment_table",
            evidence_type="official_docs_deployment",
            text=text,
        )
        if tables:
            return tables
    for block in _fenced_code_blocks(text):
        if _looks_like_solidity(block):
            tables = extract_solidity_deployment_table(block, source_url=source_url)
            if tables:
                return _retag_tables(tables, source_input_type="docs_markdown_deployment_table", evidence_type="official_docs_deployment")
        tables = json_deployment_tables(
            block,
            source_url=source_url,
            source_input_type="docs_markdown_deployment_table",
            evidence_type="official_docs_deployment",
        )
        if not tables:
            tables = yaml_deployment_tables(
                block,
                source_url=source_url,
                source_input_type="docs_markdown_deployment_table",
                evidence_type="official_docs_deployment",
            )
        if tables:
            return tables
    return []


def _retag_tables(tables: list[dict], *, source_input_type: str, evidence_type: str) -> list[dict]:
    retagged = []
    for table in tables:
        metadata = dict(table.get("metadata") or {})
        metadata["source_input_type"] = source_input_type
        metadata["evidence_type"] = evidence_type
        rows = []
        for row in table.get("rows", []):
            row = dict(row)
            row["Evidence Type"] = evidence_type
            rows.append(row)
        retagged.append({**table, "metadata": metadata, "rows": rows})
    return retagged


def _looks_like_html(text: str) -> bool:
    sample = text[:2048].lower()
    return "<table" in sample or "<html" in sample or "<!doctype html" in sample


def _looks_like_solidity(text: str) -> bool:
    sample = text[:8192]
    return "pragma solidity" in sample or bool(re.search(r"\b(?:address|I[A-Za-z0-9_]+)\s+(?:public\s+|internal\s+|private\s+|external\s+)?constant\s+[A-Z0-9_]+\s*=", sample))


def _fenced_code_blocks(text: str) -> Iterable[str]:
    for match in re.finditer(r"```[A-Za-z0-9_-]*\s*\n(.*?)```", text, flags=re.DOTALL):
        yield match.group(1)


async def fetch_url_bytes(source_url: str) -> tuple[bytes, str | None, str | None]:
    async with httpx.AsyncClient(timeout=settings.source_fetch_timeout_seconds, follow_redirects=True) as client:
        response = await client.get(source_url)
        response.raise_for_status()
        content = response.content[: settings.source_fetch_max_bytes]
        return content, str(response.url), response.headers.get("content-type")


def _parsed(
    artifact: SourceArtifact,
    fingerprint: SourceFingerprint,
    *,
    document_text: str,
    document_title: str | None,
    metadata: dict,
    table_preview: list[dict],
    candidates: list[CandidatePreview],
    warnings: list[str] | None = None,
) -> ParsedSource:
    detected_columns = [ColumnMappingService.detected_columns([str(header) for header in table.get("headers", [])]) for table in table_preview]
    metadata = {
        **metadata,
        "detected_columns": detected_columns,
        "document_sha256": hashlib.sha256(document_text.encode("utf-8")).hexdigest(),
    }
    evidence_preview = [_evidence_preview(candidate, fingerprint) for candidate in candidates[: settings.preview_candidate_limit]]
    return ParsedSource(
        document_text=document_text,
        document_title=document_title,
        content_type=artifact.content_type,
        metadata=metadata,
        table_preview=table_preview[:10],
        candidates=_dedupe_candidates(candidates),
        evidence_preview=evidence_preview,
        warnings=warnings or [],
        fatal_errors=[],
    )


def _csv_tables(raw_content: bytes) -> list[dict]:
    text = _decode(raw_content)
    reader = csv.DictReader(StringIO(text))
    headers = list(reader.fieldnames or [])
    rows = []
    for row_number, row in enumerate(reader, start=2):
        row = {str(key): "" if value is None else str(value) for key, value in row.items()}
        row["_row_number"] = row_number
        rows.append(row)
    return [{"name": "csv_registry", "headers": headers, "rows": rows, "start_line": 2}]


def _xlsx_tables(raw_content: bytes) -> tuple[list[dict], list[str], list[str], list[str]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], [], [], ["openpyxl_unavailable"]
    try:
        workbook = load_workbook(BytesIO(raw_content), read_only=True, data_only=True)
    except Exception:
        return [], [], [], ["xlsx_read_failed"]

    tables: list[dict] = []
    parsed_sheets: list[str] = []
    skipped_sheets: list[str] = []
    for sheet in workbook.worksheets:
        if _skip_xlsx_sheet(sheet.title):
            skipped_sheets.append(sheet.title)
            continue
        rows: list[tuple[int, list[str]]] = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append((row_number, values))
        if not rows:
            skipped_sheets.append(sheet.title)
            continue
        table = _table_from_rows(sheet.title, rows)
        table["sheet_name"] = sheet.title
        parsed_sheets.append(sheet.title)
        tables.append(table)
    return tables, parsed_sheets, skipped_sheets, []


def _table_from_rows(name: str, rows: list[tuple[int, list[str]]]) -> dict:
    first_row_number, header_values = rows[0]
    headers = [str(cell).strip() for cell in header_values]
    if not any(headers):
        headers = [f"column_{index + 1}" for index in range(max(len(values) for _, values in rows))]
    data_rows = rows[1:] if _looks_like_header(headers) else rows
    dict_rows = []
    for row_number, values in data_rows:
        padded = values + [""] * max(0, len(headers) - len(values))
        row_dict = {headers[index]: padded[index] for index in range(len(headers))}
        row_dict["_row_number"] = row_number
        dict_rows.append(row_dict)
    return {"name": name, "headers": headers, "rows": dict_rows, "start_line": first_row_number + 1}


def _markdown_tables(text: str) -> list[dict]:
    lines = text.splitlines()
    tables = []
    idx = 0
    while idx + 1 < len(lines):
        if "|" not in lines[idx] or not re.match(r"^\s*\|?[\s|:-]+\|?\s*$", lines[idx + 1]):
            idx += 1
            continue
        headers = [cell.strip() for cell in lines[idx].strip().strip("|").split("|")]
        rows = []
        idx += 2
        row_number = idx + 1
        while idx < len(lines) and "|" in lines[idx]:
            values = [cell.strip() for cell in lines[idx].strip().strip("|").split("|")]
            row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
            row["_row_number"] = row_number
            rows.append(row)
            idx += 1
            row_number += 1
        tables.append({"name": f"markdown_table_{len(tables) + 1}", "headers": headers, "rows": rows, "start_line": row_number})
    return tables


def _parse_hacken_audited_wallet_rows(text: str) -> list[dict]:
    lines = [(line_number, line.strip()) for line_number, line in enumerate(text.splitlines(), start=1)]
    section = _selected_audited_wallet_section(text, lines)
    start_index = section["start_index"]
    if start_index is None:
        return []

    entity_name = _detect_pdf_entity(text)
    rows = _parse_line_hacken_wallet_rows(lines, section, entity_name)
    if rows:
        return [
            _hacken_wallet_table(
                rows,
                rows[0]["_row_number"],
                "hacken_audited_wallets_line",
                {
                    "parser_stop_marker": section.get("parser_stop_marker"),
                    "audited_wallet_section_page_or_line": section.get("audited_wallet_section_page_or_line"),
                    "rejected_audited_wallet_heading_count": section.get("rejected_audited_wallet_heading_count", 0),
                    "selected_audited_wallet_heading_reason": section.get("selected_audited_wallet_heading_reason"),
                },
            )
        ]

    if section["heading_found"] and section["header_found"]:
        compact_rows, compact_metadata = _parse_compact_hacken_wallet_rows_with_metadata(text, section)
        if compact_rows:
            return [
                _hacken_wallet_table(
                    compact_rows,
                    start_index + 1,
                    "hacken_audited_wallets_compact",
                    {
                        **compact_metadata,
                        "audited_wallet_section_page_or_line": section.get("audited_wallet_section_page_or_line"),
                        "rejected_audited_wallet_heading_count": section.get("rejected_audited_wallet_heading_count", 0),
                        "selected_audited_wallet_heading_reason": section.get("selected_audited_wallet_heading_reason"),
                    },
                )
            ]
    return []


def _parse_line_hacken_wallet_rows(lines: list[tuple[int, str]], section: dict, entity_name: str | None) -> list[dict]:
    start_index = section["start_index"]
    if start_index is None:
        return []
    rows: list[dict] = []
    index = start_index
    while index < len(lines):
        line_number, line = lines[index]
        if _is_hacken_stop_line(line):
            break
        if not line or _is_footer_only_line(line):
            index += 1
            continue
        network_match = _match_known_network_at(lines, index)
        if network_match is None:
            index += 1
            continue

        network, remainder, next_index = network_match
        pieces: list[str] = [remainder] if remainder else []
        index = next_index
        while index < len(lines):
            next_line = lines[index][1] if index < len(lines) else None
            if _address_from_network_pieces(network, pieces) and not _should_continue_wrapped_address(network, pieces, next_line):
                break
            _next_number, next_line = lines[index]
            if _is_hacken_stop_line(next_line) or _is_footer_only_line(next_line) or _match_known_network_at(lines, index) is not None:
                break
            if not next_line:
                index += 1
                continue
            if _looks_like_address_continuation(next_line):
                pieces.append(next_line)
                index += 1
                continue
            break

        address = _address_from_network_pieces(network, pieces)
        if not address:
            continue
        rows.append(
            {
                "Entity": entity_name,
                "Network": network,
                "Address": address,
                "Role": "audited wallet",
                "Evidence Type": "audited_wallet",
                "Confidence": "85",
                "_row_number": line_number,
            }
        )

    return rows


def _hacken_wallet_table(rows: list[dict], start_line: int, parser: str, metadata: dict | None = None) -> dict:
    return {
        "name": "audited_wallets",
        "headers": ["Entity", "Network", "Address", "Role", "Evidence Type", "Confidence"],
        "rows": rows,
        "start_line": start_line,
        "metadata": {"parser": parser, **(metadata or {})},
    }


def _parse_hacken_layout_wallet_rows(raw_content: bytes, text: str, entity_name: str | None) -> list[dict]:
    try:
        import pdfplumber
    except Exception:
        return []

    rows: list[dict] = []
    in_section = False
    stop_marker: str | None = None
    section_line: int | None = None
    selected_reason: str | None = None
    try:
        with pdfplumber.open(BytesIO(raw_content)) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False) or []
                if not words:
                    continue
                page_text = _normalize_pdf_text(page.extract_text(x_tolerance=1, y_tolerance=3) or "")
                if not in_section:
                    if not _layout_page_has_real_audited_wallet_header(words, page_text):
                        continue
                    in_section = True
                    section_line = page_number
                    selected_reason = f"layout_header_page:{page_number}"

                page_stop = _layout_stop_marker(words)
                if page_stop:
                    stop_marker = page_stop[1]
                    words = [word for word in words if float(word["top"]) < page_stop[0]]

                rows.extend(_layout_wallet_rows_from_page(words, entity_name, page_number))
                if page_stop:
                    break
    except Exception:
        return []

    if not rows:
        return []
    return [
        _hacken_wallet_table(
            rows,
            rows[0]["_row_number"],
            "hacken_audited_wallets_layout",
            {
                "parser_stop_marker": stop_marker,
                "audited_wallet_section_page_or_line": section_line,
                "rejected_audited_wallet_heading_count": 0,
                "selected_audited_wallet_heading_reason": selected_reason,
            },
        )
    ]


def _layout_page_has_real_audited_wallet_header(words: list[dict], page_text: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", " ", page_text.lower())
    if "table of contents" in normalized:
        return False
    compact = re.sub(r"[^a-z0-9]+", "", page_text.lower())
    if "auditedwallets" not in compact:
        return False
    network_top = None
    address_top = None
    for word in words:
        value = str(word.get("text") or "").strip().lower()
        if value == "network":
            network_top = float(word["top"])
        if value == "address":
            address_top = float(word["top"])
    return network_top is not None and address_top is not None and abs(network_top - address_top) <= 8


def _layout_stop_marker(words: list[dict]) -> tuple[float, str] | None:
    lines = _layout_lines(words)
    for line in lines:
        normalized = re.sub(r"\s+", " ", line["text"].strip())
        lower = normalized.lower()
        if "collateral ratios" in lower:
            return float(line["top"]), "Collateral Ratios" if "Ratios" in normalized else "Collateral ratios"
        if lower in HACKEN_STOP_MARKERS:
            return float(line["top"]), normalized
    return None


def _layout_wallet_rows_from_page(words: list[dict], entity_name: str | None, page_number: int) -> list[dict]:
    content_words = [
        word
        for word in words
        if 115 <= float(word["top"]) <= 790
        and not _is_footer_only_line(str(word.get("text") or ""))
    ]
    lines = _layout_lines(content_words)
    data_lines = []
    for line in lines:
        text = line["text"].strip()
        if not text or _is_hacken_stop_line(text) or _is_footer_only_line(text):
            continue
        if _is_audited_wallet_heading(text) or _line_has_network_address_header(text):
            continue
        if text.lower() in {"network", "address"}:
            continue
        left = " ".join(word["text"] for word in line["words"] if float(word["x0"]) < 145).strip()
        right = " ".join(word["text"] for word in line["words"] if float(word["x0"]) >= 145).strip()
        data_lines.append({"top": float(line["top"]), "left": left, "right": right, "text": text})

    anchors: list[dict] = []
    index = 0
    while index < len(data_lines):
        line = data_lines[index]
        if not line["left"]:
            index += 1
            continue
        match = _match_known_network_prefix(line["left"])
        consumed = 1
        if match and not match[1] and index + 1 < len(data_lines) and data_lines[index + 1]["left"]:
            combined = f"{line['left']} {data_lines[index + 1]['left']}"
            combined_match = _match_known_network_prefix(combined)
            if combined_match and combined_match[0] != match[0]:
                match = combined_match
                consumed = 2
        if match:
            network, remainder = match
            top_values = [data_lines[index + offset]["top"] for offset in range(consumed)]
            anchors.append({"network": network, "remainder": remainder, "top": sum(top_values) / len(top_values), "line_index": index, "consumed": consumed})
            index += consumed
            continue
        index += 1

    rows: list[dict] = []
    for anchor_index, anchor in enumerate(anchors):
        previous_top = anchors[anchor_index - 1]["top"] if anchor_index else None
        next_top = anchors[anchor_index + 1]["top"] if anchor_index + 1 < len(anchors) else None
        start = ((previous_top + anchor["top"]) / 2) if previous_top is not None else anchor["top"] - 18
        end = ((anchor["top"] + next_top) / 2) if next_top is not None else anchor["top"] + 24
        pieces: list[str] = [anchor["remainder"]] if anchor["remainder"] else []
        for line in data_lines:
            if start <= line["top"] < end and line["right"]:
                pieces.append(line["right"])
        address = _address_from_network_pieces(anchor["network"], pieces)
        if not address:
            continue
        rows.append(
            {
                "Entity": entity_name,
                "Network": anchor["network"],
                "Address": address,
                "Role": "audited wallet",
                "Evidence Type": "audited_wallet",
                "Confidence": "85",
                "_row_number": page_number * 1000 + len(rows) + 1,
            }
        )
    return rows


def _layout_lines(words: list[dict]) -> list[dict]:
    sorted_words = sorted(words, key=lambda word: (float(word["top"]), float(word["x0"])))
    lines: list[dict] = []
    for word in sorted_words:
        top = float(word["top"])
        if not lines or abs(float(lines[-1]["top"]) - top) > 4:
            lines.append({"top": top, "words": [word]})
        else:
            lines[-1]["words"].append(word)
    for line in lines:
        line["words"].sort(key=lambda word: float(word["x0"]))
        line["text"] = " ".join(str(word.get("text") or "") for word in line["words"])
    return lines


def _parse_compact_hacken_wallet_rows(text: str) -> list[dict]:
    rows, _metadata = _parse_compact_hacken_wallet_rows_with_metadata(text)
    return rows


def _parse_compact_hacken_wallet_rows_with_metadata(text: str, section: dict | None = None) -> tuple[list[dict], dict]:
    normalized = _normalize_pdf_text(text)
    lines = [(line_number, line.strip()) for line_number, line in enumerate(normalized.splitlines(), start=1)]
    section = section or _selected_audited_wallet_section(normalized, lines)
    if not section.get("heading_found") or not section.get("header_found"):
        return [], {}
    content = _compact_section_text(normalized, lines, section)
    stop = _compact_hacken_stop(content)
    parser_stop_marker = None
    if stop is not None:
        stop_index, parser_stop_marker = stop
        content = content[:stop_index]
    compact = _compact_hacken_wallet_content(content)
    if not compact:
        return [], {"parser_stop_marker": parser_stop_marker}

    entity_name = _detect_pdf_entity(text)
    rows: list[dict] = []
    index = 0
    while index < len(compact):
        network_match = _compact_network_at(compact, index)
        if network_match is None:
            index += 1
            continue
        network, network_end = network_match
        address = _compact_address_for_network(compact, network_end, network)
        if not address:
            index = max(network_end, index + 1)
            continue
        rows.append(
            {
                "Entity": entity_name,
                "Network": network,
                "Address": address,
                "Role": "audited wallet",
                "Evidence Type": "audited_wallet",
                "Confidence": "85",
                "_row_number": 1,
            }
        )
        index = network_end + len(address)
    return rows, {"parser_stop_marker": parser_stop_marker}


def _compact_section_text(text: str, lines: list[tuple[int, str]], section: dict) -> str:
    heading_index = section.get("heading_index")
    start_index = section.get("start_index")
    if heading_index is None or start_index is None:
        return ""
    heading_line = lines[heading_index][1] if heading_index < len(lines) else ""
    header_match = re.search(r"network\s*address", heading_line, flags=re.IGNORECASE)
    if not header_match:
        compact_heading_line = re.sub(r"\s+", "", heading_line)
        compact_header = re.search(r"networkaddress", compact_heading_line, flags=re.IGNORECASE)
        if compact_header:
            raw_header_match = re.search(r"network\s*address", heading_line, flags=re.IGNORECASE)
            return heading_line[raw_header_match.end() :] if raw_header_match else re.sub(r"^.*?Network\s*Address", "", heading_line, flags=re.IGNORECASE)
    if header_match:
        return heading_line[header_match.end() :] + "\n" + "\n".join(line for _line_number, line in lines[heading_index + 1 :])
    return "\n".join(line for _line_number, line in lines[start_index:])


def _compact_hacken_stop(content: str) -> tuple[int, str] | None:
    patterns = [
        (r"collateral\s*ratios", "Collateral ratios"),
        (r"team\s*composition", "Team Composition"),
        (r"conclusion", "Conclusion"),
        (r"disclaimers", "Disclaimers"),
    ]
    matches = [(match.start(), label) for pattern, label in patterns if (match := re.search(pattern, content, flags=re.IGNORECASE))]
    return min(matches, key=lambda item: item[0]) if matches else None


def _compact_hacken_stop_index(content: str) -> int | None:
    stop = _compact_hacken_stop(content)
    return stop[0] if stop else None


def _compact_hacken_wallet_content(content: str) -> str:
    compact = re.sub(r"\s+", "", content)
    network_pattern = "|".join(re.escape(token) for token, _canonical in COMPACT_HACKEN_NETWORKS)
    noise_patterns = [
        r"Hacken'?s(?:BYBIT|MEXC|KuCoin)ProofofReserves?.{0,160}?Page\d+",
        rf"HackenOUParda4.*?(?={network_pattern}|$)",
        rf"Tallinn10151HarjuMaakond.*?(?={network_pattern}|$)",
        rf"EestiKesklinna,?Estonia.*?(?={network_pattern}|$)",
        r"Page\d+",
    ]
    for pattern in noise_patterns:
        compact = re.sub(pattern, "", compact, flags=re.IGNORECASE)
    return compact


def _compact_network_at(compact: str, index: int) -> tuple[str, int] | None:
    for token, canonical in COMPACT_HACKEN_NETWORKS:
        if compact[index : index + len(token)].lower() == token.lower():
            return canonical, index + len(token)
    return None


def _compact_address_for_network(compact: str, index: int, network: str) -> str | None:
    normalized_network = NetworkNormalizer.normalize(network)
    if normalized_network.chain_guess in {"aptos", "sui", "starknet"}:
        return _compact_0x_address(compact, index, normalized_network, min_hex=40, max_hex=64)
    if normalized_network.chain_guess == "evm":
        return _compact_0x_address(compact, index, normalized_network, min_hex=40, max_hex=40)
    for regex in _address_regexes_for_network(normalized_network):
        address = _compact_regex_address(compact, index, normalized_network, regex)
        if address:
            return address
    return None


def _compact_regex_address(compact: str, index: int, network, regex: re.Pattern) -> str | None:
    for end in range(min(len(compact), index + 96), index + 2, -1):
        candidate = compact[index:end]
        if not regex.fullmatch(candidate):
            continue
        if not _compact_boundary_after_address(compact, end):
            continue
        return candidate
    return None


def _compact_0x_address(compact: str, index: int, network, *, min_hex: int, max_hex: int) -> str | None:
    if compact[index : index + 2].lower() != "0x":
        return None
    for hex_length in range(max_hex, min_hex - 1, -1):
        end = index + 2 + hex_length
        candidate = compact[index:end]
        if not re.fullmatch(r"0x[a-fA-F0-9]+", candidate):
            continue
        if not _compact_boundary_after_address(compact, end):
            continue
        if _valid_address_for_network(candidate, network):
            return candidate
    return None


def _compact_boundary_after_address(compact: str, index: int) -> bool:
    if index >= len(compact):
        return True
    if _compact_network_at(compact, index) is not None:
        return True
    return _compact_hacken_stop_index(compact[index : index + 80]) == 0


def _pdf_audited_wallet_diagnostics(text: str, tables: list[dict]) -> dict:
    lines = [(line_number, line.strip()) for line_number, line in enumerate(text.splitlines(), start=1)]
    section = _audited_wallet_section_state(lines)
    heading_index = section["heading_index"]
    excerpt: list[str] = []
    if heading_index is not None:
        start = max(0, heading_index - 5)
        end = min(len(lines), heading_index + 15)
        excerpt = [line for _line_number, line in lines[start:end]]
    rows_detected = sum(len(table.get("rows", [])) for table in tables)
    start_index = section["start_index"]
    return {
        "pdf_text_line_count": len(lines),
        "audited_wallet_heading_found": bool(section["heading_found"]),
        "network_address_header_found": bool(section["header_found"]),
        "audited_wallet_start_line": lines[start_index][0] if isinstance(start_index, int) and start_index < len(lines) else None,
        "audited_wallet_rows_detected": rows_detected,
        "pdf_parser_mode": "hacken_audited_wallet_table" if rows_detected else "pdf_text_fallback",
        "pdf_text_excerpt_around_audited_wallet": excerpt[:20],
    }


def _hacken_pdf_parser_mode(tables: list[dict]) -> str:
    parser = (tables[0].get("metadata") or {}).get("parser") if tables else None
    if parser == "hacken_audited_wallets_layout":
        return "hacken_audited_wallet_layout_table"
    if parser == "hacken_audited_wallets_compact":
        return "hacken_audited_wallet_compact_table"
    if parser == "hacken_audited_wallets_line":
        return "hacken_audited_wallet_line_table"
    return "hacken_audited_wallet_table"


def _hacken_wallet_preview_metadata(tables: list[dict], candidates: list[CandidatePreview]) -> dict:
    raw_rows = [row for table in tables for row in table.get("rows", []) if isinstance(row, dict)]
    candidate_network_counts: dict[str, int] = {}
    for candidate in candidates:
        key = candidate.source_network or "unknown"
        candidate_network_counts[key] = candidate_network_counts.get(key, 0) + 1

    network_counts: dict[str, int] = {}
    unsupported_network_counts: dict[str, int] = {}
    unsupported_address_families: set[str] = set()
    for row in raw_rows:
        network = str(row.get("Network") or "unknown")
        network_counts[network] = network_counts.get(network, 0) + 1
        candidate_count = candidate_network_counts.get(network, 0)
        if candidate_count >= network_counts[network]:
            continue
        unsupported_network_counts[network] = unsupported_network_counts.get(network, 0) + 1
        normalized = NetworkNormalizer.normalize(network)
        unsupported_address_families.add(normalized.chain_guess or normalized.canonical_chain or "unknown")

    table_metadata = tables[0].get("metadata") if tables else {}
    skipped_count = max(0, len(raw_rows) - len(candidates))
    return {
        "total_wallet_rows_detected": len(raw_rows),
        "raw_wallet_rows_detected": len(raw_rows),
        "candidate_rows_created": len(candidates),
        "skipped_raw_rows_count": skipped_count,
        "network_counts": network_counts,
        "candidate_network_counts": candidate_network_counts,
        "unsupported_network_counts": unsupported_network_counts,
        "unsupported_address_families": sorted(unsupported_address_families),
        "parser_stop_marker": (table_metadata or {}).get("parser_stop_marker"),
        "audited_wallet_section_page_or_line": (table_metadata or {}).get("audited_wallet_section_page_or_line"),
        "rejected_audited_wallet_heading_count": (table_metadata or {}).get("rejected_audited_wallet_heading_count", 0),
        "selected_audited_wallet_heading_reason": (table_metadata or {}).get("selected_audited_wallet_heading_reason"),
    }


def _audited_wallet_section_start(lines: list[tuple[int, str]]) -> int | None:
    return _audited_wallet_section_state(lines)["start_index"]


def _audited_wallet_section_state(lines: list[tuple[int, str]]) -> dict:
    return _selected_audited_wallet_section("\n".join(line for _line_number, line in lines), lines)


def _selected_audited_wallet_section(text: str, lines: list[tuple[int, str]] | None = None) -> dict:
    lines = lines or [(line_number, line.strip()) for line_number, line in enumerate(text.splitlines(), start=1)]
    candidates = _audited_wallet_section_candidates(lines)
    if not candidates:
        return {"heading_found": False, "header_found": False, "heading_index": None, "start_index": None}

    entity_name = _detect_pdf_entity(text)
    scored: list[tuple[int, int, dict]] = []
    rejected = 0
    for candidate in candidates:
        line_rows = _parse_line_hacken_wallet_rows(lines, candidate, entity_name)
        compact_rows, _compact_metadata = _parse_compact_hacken_wallet_rows_with_metadata(text, candidate) if candidate.get("header_found") else ([], {})
        row_count = max(len(line_rows), len(compact_rows))
        toc_marker_count = _toc_marker_count(lines, int(candidate["heading_index"]))
        is_toc = toc_marker_count >= 3
        candidate["parsed_row_count"] = row_count
        candidate["toc_marker_count"] = toc_marker_count
        if (not candidate.get("header_found") and row_count < 3) or is_toc:
            rejected += 1
            continue
        score = row_count * 100 + (20 if candidate.get("header_found") else 0) - toc_marker_count * 10
        scored.append((score, row_count, candidate))

    if not scored:
        first = candidates[0]
        first["rejected_audited_wallet_heading_count"] = max(0, len(candidates) - 1)
        first["selected_audited_wallet_heading_reason"] = "no_structured_wallet_section_selected"
        return first

    scored.sort(key=lambda item: (item[0], item[1], -int(item[2]["heading_index"])), reverse=True)
    selected = dict(scored[0][2])
    selected["rejected_audited_wallet_heading_count"] = rejected + max(0, len(candidates) - len(scored) - rejected)
    selected["selected_audited_wallet_heading_reason"] = f"highest_parsed_row_count:{scored[0][1]}"
    return selected


def _audited_wallet_section_candidates(lines: list[tuple[int, str]]) -> list[dict]:
    candidates: list[dict] = []
    for index, (_line_number, line) in enumerate(lines):
        if not _is_audited_wallet_heading(line):
            continue
        if _line_has_network_address_header(line):
            candidates.append(_audited_wallet_section_candidate(lines, index, True, index))
            continue
        network_line: int | None = None
        address_line: int | None = None
        selected_header_index: int | None = None
        for header_index in range(index + 1, min(index + 10, len(lines))):
            header_line = lines[header_index][1]
            header = re.sub(r"\s+", " ", header_line.lower()).strip()
            if "network" in header and network_line is None:
                network_line = header_index
            if "address" in header and address_line is None:
                address_line = header_index
            if (network_line is not None and address_line is not None) or _line_has_network_address_header(header_line):
                selected_header_index = max(value for value in [network_line, address_line, header_index] if value is not None)
                break
        if selected_header_index is not None:
            candidates.append(_audited_wallet_section_candidate(lines, index, True, selected_header_index + 1))
        else:
            candidates.append(_audited_wallet_section_candidate(lines, index, False, index + 1))
    return candidates


def _audited_wallet_section_candidate(lines: list[tuple[int, str]], heading_index: int, header_found: bool, start_index: int) -> dict:
    stop_marker = None
    for _line_number, line in lines[start_index:]:
        if _is_hacken_stop_line(line):
            stop_marker = re.sub(r"\s+", " ", line.strip()) or None
            break
    return {
        "heading_found": True,
        "header_found": header_found,
        "heading_index": heading_index,
        "start_index": start_index,
        "audited_wallet_section_page_or_line": lines[heading_index][0],
        "parser_stop_marker": stop_marker,
    }


def _toc_marker_count(lines: list[tuple[int, str]], heading_index: int) -> int:
    nearby = " ".join(line for _line_number, line in lines[max(0, heading_index - 8) : min(len(lines), heading_index + 22)])
    normalized = re.sub(r"[^a-z0-9]+", " ", nearby.lower())
    return sum(1 for marker in HACKEN_TOC_MARKERS if marker in normalized)


def _is_audited_wallet_heading(line: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", " ", line.lower()).strip()
    compact = re.sub(r"[^a-z0-9]+", "", line.lower())
    return "audited wallets" in normalized or "auditedwallets" in compact


def _line_has_network_address_header(line: str) -> bool:
    compact = re.sub(r"[^a-z0-9]+", "", line.lower())
    return "networkaddress" in compact


def _is_hacken_stop_line(line: str) -> bool:
    normalized = re.sub(r"\s+", " ", line.strip().lower())
    return any(marker in normalized for marker in HACKEN_STOP_MARKERS)


def _is_footer_only_line(line: str) -> bool:
    normalized = line.strip()
    if not normalized:
        return True
    if re.fullmatch(r"\d{1,3}", normalized):
        return True
    if re.fullmatch(r"page\s+\d+(?:\s+of\s+\d+)?", normalized, flags=re.IGNORECASE):
        return True
    lower = normalized.lower()
    if "bybit proof of reserve" in lower or "bybit proof of reserves" in lower:
        return True
    if "mexc proof of reserve" in lower or "mexc proof of reserves" in lower:
        return True
    if "kucoin proof of reserve" in lower or "kucoin proof of reserves" in lower:
        return True
    if "hacken" in lower and "proof of reserve" in lower:
        return True
    if lower.startswith(("hacken ou", "parda 4", "tallinn 10151", "eesti kesklinna")):
        return True
    return normalized.lower() in {"hacken", "proof of reserves audit report", "bybit proof of reserves audit report", "mexc proof of reserves audit report", "kucoin proof of reserves audit report"}


def _match_known_network_at(lines: list[tuple[int, str]], index: int) -> tuple[str, str, int] | None:
    line = lines[index][1]
    direct = _match_known_network_prefix(line)
    if direct is not None:
        network, remainder = direct
        if not remainder and index + 1 < len(lines):
            combined_line = f"{line} {lines[index + 1][1]}".strip()
            combined = _match_known_network_prefix(combined_line)
            if combined is not None and (combined[0] != network or combined[1] != lines[index + 1][1]):
                return combined[0], combined[1], index + 2
        return network, remainder, index + 1

    if index + 1 < len(lines):
        combined_line = f"{line} {lines[index + 1][1]}".strip()
        combined = _match_known_network_prefix(combined_line)
        if combined is not None and len(combined[0].split()) > 1:
            return combined[0], combined[1], index + 2
    return None


def _match_known_network_prefix(line: str) -> tuple[str, str] | None:
    normalized = re.sub(r"\s+", " ", line.strip().replace("/", " "))
    for network in HACKEN_NETWORK_LABELS:
        pattern = re.compile(rf"^{re.escape(network)}(?=$|\s|:|-)", flags=re.IGNORECASE)
        match = pattern.match(normalized)
        if match:
            return HACKEN_NETWORK_ALIASES.get(network, network), normalized[match.end() :].strip(" :-\t")
    return None


def _looks_like_address_continuation(line: str) -> bool:
    stripped = re.sub(r"\s+", "", line.strip())
    if not stripped:
        return False
    if _match_known_network_prefix(line) is not None:
        return False
    return bool(re.fullmatch(r"[A-Fa-f0-9]{4,64}", stripped) or re.fullmatch(r"[A-Za-z0-9_-]{8,96}", stripped))


def _address_from_network_pieces(network: str, pieces: list[str]) -> str | None:
    compact = "".join(re.sub(r"\s+", "", piece.strip()) for piece in pieces if piece and piece.strip())
    if not compact:
        return None
    normalized_network = NetworkNormalizer.normalize(network)
    if normalized_network.chain_guess == "substrate":
        match = SUBSTRATE_RE.search(compact)
        if match:
            address = match.group(0)
            return address if _valid_address_for_network(address, normalized_network) else None
    if normalized_network.chain_guess in {"aptos", "sui", "starknet"}:
        match = re.search(r"0x[a-fA-F0-9]{40,64}", compact)
        if match:
            address = match.group(0)
            return address if _valid_address_for_network(address, normalized_network) else None
    if normalized_network.chain_guess == "evm":
        match = re.search(r"0x[a-fA-F0-9]{40}(?![a-fA-F0-9])", compact)
        if match:
            address = match.group(0)
            return address if _valid_address_for_network(address, normalized_network) else None
    for regex in _address_regexes_for_network(normalized_network):
        for match in regex.finditer(compact):
            address = match.group(0)
            if _valid_address_for_network(address, normalized_network):
                return address
    match = ADDRESS_RE.search(compact)
    if match and _valid_address_for_network(match.group(0), normalized_network):
        return match.group(0)
    return None


def _should_continue_wrapped_address(network: str, pieces: list[str], next_line: str | None) -> bool:
    if not next_line:
        return False
    if _is_hacken_stop_line(next_line) or _is_footer_only_line(next_line) or _match_known_network_prefix(next_line):
        return False
    normalized_network = NetworkNormalizer.normalize(network)
    if normalized_network.chain_guess == "evm":
        return False
    compact = "".join(re.sub(r"\s+", "", piece.strip()) for piece in pieces if piece and piece.strip())
    continuation = re.sub(r"\s+", "", next_line.strip())
    if not _looks_like_address_continuation(continuation):
        return False
    if normalized_network.chain_guess in {"aptos", "sui", "starknet"}:
        match = re.search(r"0x[a-fA-F0-9]{1,63}$", compact)
        return bool(match and len(match.group(0)) - 2 + len(continuation) <= 64 and re.fullmatch(r"[a-fA-F0-9]{1,64}", continuation))
    if normalized_network.chain_guess in {"btc", "bitcoin", "algorand", "solana", "substrate", "near", "cosmos", "ton", "tezos", "litecoin"}:
        return True
    return not _address_from_network_pieces(network, pieces)


def _address_regexes_for_network(network) -> list[re.Pattern]:
    regexes: list[re.Pattern] = []
    if network.canonical_chain == "xdc":
        regexes.extend([XDC_RE, EVM_RE])
    if network.chain_guess == "hedera":
        regexes.append(HEDERA_RE)
    if network.chain_guess == "solana":
        regexes.append(SOLANA_RE)
    if network.chain_guess == "substrate":
        regexes.append(SUBSTRATE_RE)
    if network.canonical_chain == "vaulta":
        regexes.append(VAULTA_RE)
    if network.chain_guess == "algorand":
        regexes.append(ALGORAND_RE)
    if network.chain_guess == "near":
        regexes.append(NEAR_RE)
    if network.canonical_chain == "noble":
        regexes.append(NOBLE_RE)
    if network.chain_guess == "tezos":
        regexes.append(TEZOS_RE)
    regexes.extend([BTC_RE, LTC_RE, TRON_RE, TON_RE, XRP_RE, COSMOS_RE, DYDX_RE, DOGE_RE, AVALANCHE_X_RE])
    return regexes


def _extract_table_candidates(
    artifact: SourceArtifact,
    fingerprint: SourceFingerprint,
    tables: list[dict],
    *,
    default_source_input_type: str,
) -> list[CandidatePreview]:
    candidates: list[CandidatePreview] = []
    for table in tables:
        table_meta = table.get("metadata") or {}
        table_source_input_type = table_meta.get("source_input_type") or default_source_input_type
        table_evidence_type = table_meta.get("evidence_type")
        mapping = ColumnMappingService.map_headers([str(header) for header in table.get("headers", [])])
        for row in table.get("rows", []):
            if not isinstance(row, dict) or not any(str(value).strip() for value in row.values()):
                continue
            raw_network = mapping.get(row, "chain")
            role_hint = mapping.get(row, "role")
            source_url, source_file_name = _source_reference_from_row(mapping.get(row, "source_url"), artifact.source_url)
            row_number = _int_or_none(mapping.get(row, "source_row")) or _int_or_none(str(row.get("_row_number")))
            base_reference = _row_reference(
                row=row,
                table=table,
                mapping=mapping,
                source_file_name=source_file_name,
                default_source_input_type=table_source_input_type,
            )
            for address_kind, address, role in _structured_address_fields(mapping, row, role_hint, table):
                candidate = _candidate_from_address(
                    artifact,
                    fingerprint,
                    address=address,
                    raw_network=raw_network,
                    role_hint=role_hint,
                    suggested_role=role,
                    source_url=source_url,
                    source_sheet=table.get("sheet_name"),
                    source_row=row_number,
                    source_page=_int_or_none(mapping.get(row, "source_row")) if "page" in (mapping.columns.get("source_row") or "").lower() else None,
                    table_name=table.get("name"),
                    evidence_type=mapping.get(row, "evidence_type") or table_evidence_type or "source_extraction_context",
                    source_input_type=table_source_input_type,
                    raw_reference={**base_reference, "address_column_kind": address_kind},
                )
                if candidate:
                    candidates.append(candidate)
            if any(mapping.get(row, field) for field in ("address", "deposit_address", "withdrawal_address")):
                continue
            for match in ADDRESS_RE.finditer(" ".join(_visible_row_values(row))):
                candidate = _candidate_from_address(
                    artifact,
                    fingerprint,
                    address=match.group(0),
                    raw_network=raw_network,
                    role_hint=role_hint,
                    suggested_role=_suggest_role(role_hint),
                    source_url=source_url,
                    source_sheet=table.get("sheet_name"),
                    source_row=row_number,
                    source_page=None,
                    table_name=table.get("name"),
                    evidence_type=mapping.get(row, "evidence_type") or table_evidence_type or "source_extraction_context",
                    source_input_type=table_source_input_type,
                    raw_reference={**base_reference, "address_column_kind": "row_regex"},
                )
                if candidate:
                    candidates.append(candidate)
    return candidates


def _extract_text_candidates(
    artifact: SourceArtifact,
    fingerprint: SourceFingerprint,
    text: str,
    *,
    source_input_type: str = "plain_text",
) -> list[CandidatePreview]:
    candidates: list[CandidatePreview] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        for match in ADDRESS_RE.finditer(line):
            raw_network = _network_from_line(line)
            warnings = ["no_source_network_for_text_candidate"] if source_input_type == "pdf_text_fallback" and not raw_network else []
            candidate = _candidate_from_address(
                artifact,
                fingerprint,
                address=match.group(0),
                raw_network=raw_network,
                role_hint=None,
                suggested_role=None,
                source_url=artifact.source_url,
                source_sheet=None,
                source_row=line_number,
                source_page=None,
                table_name=None,
                evidence_type="source_extraction_context",
                source_input_type=source_input_type,
                raw_reference={"raw_line": line, "line_number": line_number, "warnings": warnings},
            )
            if candidate:
                candidates.append(candidate)
    return candidates


def _candidate_from_address(
    artifact: SourceArtifact,
    fingerprint: SourceFingerprint,
    *,
    address: str,
    raw_network: str | None,
    role_hint: str | None,
    suggested_role: str | None,
    source_url: str | None,
    source_sheet: str | None,
    source_row: int | None,
    source_page: int | None,
    table_name: str | None,
    evidence_type: str,
    source_input_type: str,
    raw_reference: dict,
) -> CandidatePreview | None:
    clean = clean_wallet_address(address)
    if not clean:
        return None
    network = NetworkNormalizer.normalize(raw_network)
    if not _valid_address_for_network(clean, network):
        return None
    inferred_family = _infer_address_family(clean)
    chain_guess = network.chain_guess or inferred_family
    chain_slug = network.canonical_chain
    chain_id = network.chain_id
    warnings: list[str] = list(raw_reference.get("warnings") or [])
    if raw_network and not network.canonical_chain:
        warnings.append("unrecognized_source_network")
    if raw_reference.get("validator_public_key"):
        warnings.append("validator_public_key_metadata_only")
    if network.canonical_chain == "xdc":
        normalized = f"xdc{clean[2:].lower()}" if clean.lower().startswith("0x") else clean.lower()
    else:
        normalized = clean.lower() if clean.startswith("0x") else clean
    confidence = _confidence_from_value(raw_reference.get("confidence")) or (70 if raw_network else 45)
    entity = raw_reference.get("row_entity") or _entity_from_sheet_name(source_sheet)
    if not raw_network and source_input_type in {
        "docs_html_deployment_table",
        "docs_markdown_deployment_table",
        "github_solidity_address_book",
        "github_json_deployment_registry",
        "github_markdown_deployment_table",
    }:
        warnings.append("missing_network_context")
    return CandidatePreview(
        address=clean,
        normalized_address=normalized,
        entity_name=entity,
        source_network=raw_network,
        chain_guess=chain_guess,
        chain_slug=chain_slug,
        chain_id=chain_id,
        address_family=_address_family(chain_guess, inferred_family),
        suggested_role=suggested_role or _suggest_role(role_hint),
        confidence_initial=confidence,
        status="needs_review",
        source_type=fingerprint.final_source_type or "unknown",
        source_input_type=source_input_type,
        source_sheet=source_sheet,
        source_row=source_row,
        source_page=source_page,
        source_url=source_url,
        file_path=artifact.local_file_path,
        evidence_type=evidence_type,
        warnings=warnings,
        raw_reference={
            **raw_reference,
            "table_name": table_name,
            "original_value": address,
            "normalized_value": normalized,
            "final_source_type": fingerprint.final_source_type,
            "adapter_name": fingerprint.parser_adapter,
            "source_url": source_url,
            "file_path": artifact.local_file_path,
        },
    )


def _structured_address_fields(mapping: ColumnMapping, row: dict, role_hint: str | None, table: dict) -> list[tuple[str, str, str | None]]:
    fields: list[tuple[str, str, str | None]] = []
    address = mapping.get(row, "address")
    if address:
        fields.append(("address", address, _suggest_role(role_hint) or "cex_reserve_wallet_candidate"))
    deposit_address = mapping.get(row, "deposit_address")
    if deposit_address:
        fields.append(("deposit_address", deposit_address, "staking_deposit_wallet"))
    withdrawal_address = mapping.get(row, "withdrawal_address")
    if withdrawal_address:
        role = "staking_withdrawal_wallet" if "staking" in _normalize_sheet_name(str(table.get("sheet_name") or table.get("name") or "")) else "cex_cold_wallet"
        fields.append(("withdrawal_address", withdrawal_address, role))
    return fields


def _row_reference(*, row: dict, table: dict, mapping: ColumnMapping, source_file_name: str | None, default_source_input_type: str) -> dict:
    validator_public_key = mapping.get(row, "validator_public_key")
    deployment_details = _deployment_raw_details(row)
    raw_row = deployment_details.get("raw_row") if isinstance(deployment_details.get("raw_row"), dict) else None
    return {
        "raw_row_json": {key: value for key, value in row.items() if not str(key).startswith("_")},
        "source_input_type": default_source_input_type,
        "source_sheet": table.get("sheet_name"),
        "source_row": row.get("_row_number"),
        "source_file_name": source_file_name,
        "row_entity": mapping.get(row, "entity"),
        "row_protocol": mapping.get(row, "protocol"),
        "row_category": mapping.get(row, "category"),
        "evidence_type": mapping.get(row, "evidence_type"),
        "contract_name": deployment_details.get("contract_name") or _first_present_cell(row, "Contract Name", "contract_name", "Contract", "Name"),
        "role_source": deployment_details.get("role_source") or _first_present_cell(row, "Role", "role"),
        "column_name": deployment_details.get("column_name"),
        "line_number": deployment_details.get("line_number") or mapping.get(row, "source_row"),
        "raw_row": raw_row,
        "raw_line": raw_row.get("raw_line") if raw_row else None,
        "report_date": mapping.get(row, "report_date"),
        "audit_date": mapping.get(row, "audit_date"),
        "confidence": mapping.get(row, "confidence"),
        "notes": mapping.get(row, "notes"),
        "validator_public_key": validator_public_key,
        "warnings": ["validator_public_key_metadata_only"] if validator_public_key else [],
    }


def _deployment_raw_details(row: dict) -> dict:
    raw = _first_present_cell(row, "Raw Row JSON", "raw_row_json")
    if not raw:
        return {}
    try:
        parsed = json.loads(str(raw))
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _first_present_cell(row: dict, *keys: str) -> str | None:
    normalized = {re.sub(r"[^a-z0-9]+", "_", str(key).strip().lower()).strip("_"): key for key in row}
    for key in keys:
        actual = normalized.get(re.sub(r"[^a-z0-9]+", "_", key.strip().lower()).strip("_"))
        if actual is not None and row.get(actual) not in {None, ""}:
            return str(row.get(actual)).strip()
    return None


def clean_wallet_address(value: str | None) -> str | None:
    if not value:
        return None
    cleaned = str(value).strip().strip("`'\";,()[]{}")
    cleaned = re.sub(r"\s+", "", cleaned)
    if len(cleaned) < 8:
        return None
    return cleaned


def _valid_address_for_network(address: str, network) -> bool:
    chain_guess = network.chain_guess
    if EVM_RE.fullmatch(address):
        return chain_guess not in {"aptos", "sui", "xrp", "ton", "tron", "btc", "bitcoin"}
    if LONG_0X_RE.fullmatch(address):
        return chain_guess in {"aptos", "sui", "starknet"}
    if XRP_RE.fullmatch(address):
        return chain_guess == "xrp"
    if BTC_RE.fullmatch(address):
        return chain_guess in {"btc", "bitcoin", "litecoin", "dogecoin"}
    if LTC_RE.fullmatch(address):
        return chain_guess == "litecoin"
    if TRON_RE.fullmatch(address):
        return chain_guess == "tron"
    if TON_RE.fullmatch(address):
        return chain_guess == "ton"
    if SUBSTRATE_RE.fullmatch(address) and chain_guess == "substrate":
        return True
    if COSMOS_RE.fullmatch(address):
        return network.canonical_chain == "cosmos"
    if DYDX_RE.fullmatch(address):
        return network.canonical_chain == "dydx"
    if DOGE_RE.fullmatch(address):
        return chain_guess == "dogecoin"
    if AVALANCHE_X_RE.fullmatch(address):
        return network.canonical_chain == "avalanche-x"
    if TEZOS_RE.fullmatch(address):
        return chain_guess == "tezos"
    if SOLANA_RE.fullmatch(address):
        return chain_guess == "solana"
    if HEDERA_RE.fullmatch(address):
        return chain_guess == "hedera"
    if VAULTA_RE.fullmatch(address) and network.canonical_chain == "vaulta":
        return True
    if XDC_RE.fullmatch(address):
        return network.canonical_chain == "xdc"
    if ALGORAND_RE.fullmatch(address):
        return chain_guess == "algorand"
    if NEAR_RE.fullmatch(address):
        return chain_guess == "near"
    if NOBLE_RE.fullmatch(address):
        return network.canonical_chain == "noble"
    return False


def _suggest_role(role_hint: str | None) -> str | None:
    if not role_hint:
        return None
    text = role_hint.lower()
    if "deposit" in text and "staking" in text:
        return "staking_deposit_wallet"
    if "withdrawal" in text:
        return "staking_withdrawal_wallet" if "staking" in text else "cex_cold_wallet"
    if "cold" in text:
        return "cex_cold_wallet"
    if "hot" in text:
        return "cex_hot_wallet"
    if "reserve" in text or "por" in text or "audited" in text:
        return "cex_por_wallet"
    if "factory" in text:
        return "factory_contract"
    if "router" in text:
        return "router_contract"
    if "pool_addresses_provider" in text or "addresses_provider" in text or "address provider" in text:
        return "address_provider"
    if "oracle" in text:
        return "oracle"
    if "configurator" in text:
        return "protocol_configurator"
    if "collector" in text or "treasury" in text:
        return "treasury"
    if "nftdescriptor" in text or "nft descriptor" in text or "descriptor" in text:
        return "nft_descriptor"
    if "flow" in text or "stream" in text or "lockup" in text:
        return "protocol_contract"
    if "token" in text:
        return "token_contract"
    if "pool" in text:
        return "liquidity_pool"
    normalized = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return normalized or None


def _source_reference_from_row(value: str | None, fallback_url: str | None) -> tuple[str | None, str | None]:
    if not value:
        return fallback_url, None
    text = value.strip()
    if URL_RE.match(text):
        return text, None
    return fallback_url, text


def _evidence_preview(candidate: CandidatePreview, fingerprint: SourceFingerprint) -> dict:
    return {
        "address": candidate.address,
        "source_type": candidate.source_type,
        "final_source_type": fingerprint.final_source_type,
        "adapter_name": fingerprint.parser_adapter,
        "source_url": candidate.source_url,
        "file_path": candidate.file_path,
        "sheet_name": candidate.source_sheet,
        "row_number": candidate.source_row,
        "page_number": candidate.source_page,
        "raw_reference": candidate.raw_reference,
        "confidence_reason": "structured_network_column" if candidate.source_network else "address_pattern_fallback",
    }


def _dedupe_candidates(candidates: list[CandidatePreview]) -> list[CandidatePreview]:
    seen: set[tuple[str, str | None, int | None, str | None, int | None, str | None]] = set()
    deduped: list[CandidatePreview] = []
    for candidate in candidates:
        key = (
            candidate.normalized_address,
            candidate.chain_slug,
            candidate.chain_id,
            candidate.source_sheet,
            candidate.source_row,
            candidate.suggested_role,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(candidate)
    return deduped


def _decode(content: bytes, fallback: str = "") -> str:
    if not content:
        return fallback
    return content.decode("utf-8-sig", errors="replace")


def _normalize_pdf_text(text: str) -> str:
    if not text:
        return ""
    normalized = text.translate(PDF_LIGATURE_TRANSLATION)
    for old, new in PDF_TEXT_REPLACEMENTS:
        normalized = normalized.replace(old, new)
    normalized = normalized.replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"[\t\f\v\u00a0\u1680\u180e\u2000-\u200b\u2028\u2029\u202f\u205f\u3000]", " ", normalized)
    normalized = "\n".join(re.sub(r" {2,}", " ", line).strip() for line in normalized.split("\n"))
    return normalized


def _skip_xlsx_sheet(name: str | None) -> bool:
    if not name:
        return True
    normalized = _normalize_sheet_name(name)
    if normalized in CONTROL_SHEET_NAMES:
        return True
    if any(token in normalized for token in {"provenance", "backlog", "capture status", "schema"}):
        return True
    return normalized.startswith(("summary", "readme", "instruction"))


def _normalize_sheet_name(name: str) -> str:
    normalized = re.sub(r"[_\-/]+", " ", name.strip().lower())
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _entity_from_sheet_name(value: str | None) -> str | None:
    if not value:
        return None
    normalized = _normalize_sheet_name(value)
    if not normalized or _skip_xlsx_sheet(normalized):
        return None
    tokens = [token for token in normalized.split() if token not in SHEET_ENTITY_SUFFIX_TOKENS]
    if not tokens:
        return None
    if tokens[:2] == ["huobi", "htx"]:
        return "Huobi_HTX"
    primary = tokens[0]
    return SHEET_ENTITY_CASING.get(primary, primary.upper() if len(primary) <= 4 else primary.title())


def _network_from_line(line: str) -> str | None:
    before = ADDRESS_RE.split(line, maxsplit=1)[0]
    tokens = before.strip(" :-\t|")
    if not tokens:
        return None
    parts = re.split(r"\s{2,}|\t|\|", tokens)
    for part in reversed(parts):
        normalized = NetworkNormalizer.normalize(part)
        if normalized.canonical_chain:
            return part
    return tokens if NetworkNormalizer.normalize(tokens).canonical_chain else None


def _infer_address_family(address: str) -> str | None:
    lower = address.lower()
    if lower.startswith("0x"):
        return "evm"
    if lower.startswith("bc1") or re.match(r"^[13][a-km-zA-HJ-NP-Z1-9]{25,34}$", address):
        return "btc"
    if address.startswith("T") and len(address) == 34:
        return "tron"
    if address.startswith("r") and 25 <= len(address) <= 35:
        return "xrp"
    if address.startswith(("EQ", "UQ")):
        return "ton"
    if lower.startswith("cosmos1") or lower.startswith("dydx1"):
        return "cosmos"
    if address.startswith("D"):
        return "dogecoin"
    if lower.startswith("x-avax1"):
        return "avalanche"
    if lower.startswith("ltc1") or address[:1] in {"L", "M"}:
        return "litecoin"
    if lower.startswith("xdc"):
        return "evm"
    if re.fullmatch(r"0\.0\.\d{1,12}", address):
        return "hedera"
    if lower.endswith(".near") or re.fullmatch(r"[a-f0-9]{64}", lower):
        return "near"
    if lower.startswith("noble1"):
        return "cosmos"
    if lower.startswith("tz1"):
        return "tezos"
    if ALGORAND_RE.fullmatch(address):
        return "algorand"
    if SOLANA_RE.fullmatch(address):
        return "solana"
    return None


def _address_family(*values: str | None) -> str | None:
    for value in values:
        if value:
            return "bitcoin" if value == "btc" else value
    return None


def _confidence_from_value(value: object) -> int | None:
    if value in {None, ""}:
        return None
    text = str(value).strip().lower()
    if text in {"high", "confirmed", "strong"}:
        return 95
    if text in {"medium", "med", "review", "needs review"}:
        return 75
    if text in {"low", "weak", "candidate"}:
        return 55
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group(0))
    if number <= 1:
        number *= 100
    return max(0, min(100, int(round(number))))


def _visible_row_values(row: dict) -> list[str]:
    return [str(value) for key, value in row.items() if not str(key).startswith("_") and value not in {None, ""}]


def _tables_to_lines(tables: list[dict]) -> Iterable[str]:
    for table in tables:
        yield ",".join(str(header) for header in table.get("headers", []))
        for row in table.get("rows", []):
            if isinstance(row, dict):
                yield ",".join(_visible_row_values(row))


def _file_name(path: str | None) -> str | None:
    if not path:
        return None
    return Path(path).name


def _title_from_url(source_url: str | None) -> str | None:
    if not source_url:
        return None
    return Path(urlparse(source_url).path).name or source_url


def _detect_pdf_entity(text: str) -> str | None:
    if re.search(r"\bAuditee\s+MEXC\b", text, flags=re.IGNORECASE):
        return "MEXC"
    if re.search(r"\bAuditee\s+KuCoin\b", text, flags=re.IGNORECASE):
        return "KuCoin"
    if re.search(r"\bAuditee\s+Bybit\b", text, flags=re.IGNORECASE):
        return "Bybit"
    if re.search(r"\bHacken'?s?\s+MEXC\s+Proof\s+of\s+Reserves?\b", text, flags=re.IGNORECASE):
        return "MEXC"
    if re.search(r"\bHacken'?s?\s+KuCoin\s+Proof\s+of\s+Reserves?\b", text, flags=re.IGNORECASE):
        return "KuCoin"
    if re.search(r"\bMEXC\b", text):
        return "MEXC"
    if re.search(r"\bKUCOIN\b|\bKuCoin\b", text):
        return "KuCoin"
    if re.search(r"\bBYBIT\b", text):
        return "Bybit"
    return None


def _is_hacken_por_text(text: str, entity_name: str | None = None) -> bool:
    if not _is_proof_of_reserves_text(text):
        return False
    normalized = re.sub(r"[^a-z0-9]+", " ", text.lower())
    compact = re.sub(r"[^a-z0-9]+", "", text.lower())
    return "hacken" in normalized or "hacken" in compact


def _is_proof_of_reserves_text(text: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return "proof of reserves" in normalized or "proof of reserve" in normalized


def _first_line(text: str) -> str | None:
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            return stripped[:255]
    return None


def _int_or_none(value: str | None) -> int | None:
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _looks_like_header(headers: list[str]) -> bool:
    mapped = ColumnMappingService.map_headers(headers)
    return bool(mapped.columns) or any(re.search(r"[A-Za-z]", header) for header in headers)
